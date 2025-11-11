from flask import Flask, render_template, request, jsonify, send_file
from datetime import datetime
import sqlite3
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.lib.units import inch
import io

app = Flask(__name__)

# Configuración
DATABASE = 'inventix.db'

# Inicializar base de datos
def init_db():
    if not os.path.exists(DATABASE):
        conn = sqlite3.connect(DATABASE)
        cursor = conn.cursor()
        
        # Tabla de clientes
        cursor.execute('''CREATE TABLE clientes (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nombre TEXT NOT NULL,
            cedula TEXT UNIQUE NOT NULL,
            celular TEXT NOT NULL,
            email TEXT NOT NULL,
            direccion TEXT NOT NULL,
            fecha_registro TEXT DEFAULT CURRENT_TIMESTAMP
        )''')
        
        # Tabla de productos
        cursor.execute('''CREATE TABLE productos (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nombre TEXT NOT NULL,
            descripcion TEXT,
            precio REAL NOT NULL,
            stock INTEGER NOT NULL,
            proveedor TEXT,
            fecha_registro TEXT DEFAULT CURRENT_TIMESTAMP
        )''')
        
        # Tabla de ventas
        cursor.execute('''CREATE TABLE ventas (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            numero_factura TEXT UNIQUE NOT NULL,
            cliente_id INTEGER NOT NULL,
            total REAL NOT NULL,
            fecha TEXT NOT NULL,
            hora TEXT NOT NULL,
            FOREIGN KEY(cliente_id) REFERENCES clientes(id)
        )''')
        
        # Tabla de detalle de ventas
        cursor.execute('''CREATE TABLE detalle_ventas (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            venta_id INTEGER NOT NULL,
            producto_id INTEGER NOT NULL,
            cantidad INTEGER NOT NULL,
            precio REAL NOT NULL,
            subtotal REAL NOT NULL,
            FOREIGN KEY(venta_id) REFERENCES ventas(id),
            FOREIGN KEY(producto_id) REFERENCES productos(id)
        )''')
        
        conn.commit()
        conn.close()

# Funciones de utilidad
def get_db_connection():
    conn = sqlite3.connect(DATABASE)
    conn.row_factory = sqlite3.Row
    return conn

def obtener_numero_factura():
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute('SELECT COUNT(*) as count FROM ventas')
    count = cursor.fetchone()[0]
    conn.close()
    return f"FAC-{count + 1:05d}"

# Rutas
@app.route('/')
def index():
    conn = get_db_connection()
    cursor = conn.cursor()
    
    cursor.execute('SELECT COUNT(*) as count FROM clientes')
    total_clientes = cursor.fetchone()[0]
    
    cursor.execute('SELECT COUNT(*) as count FROM productos')
    total_productos = cursor.fetchone()[0]
    
    cursor.execute('SELECT COUNT(*) as count FROM ventas')
    total_ventas = cursor.fetchone()[0]
    
    cursor.execute('SELECT SUM(total) as total FROM ventas')
    total_vendido = cursor.fetchone()[0] or 0
    
    conn.close()
    
    return render_template('index.html', 
                          total_clientes=total_clientes,
                          total_productos=total_productos,
                          total_ventas=total_ventas,
                          total_vendido=total_vendido)

# CLIENTES
@app.route('/clientes')
def clientes():
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM clientes ORDER BY nombre')
    clientes = cursor.fetchall()
    conn.close()
    return render_template('clientes.html', clientes=clientes)

@app.route('/api/clientes', methods=['POST'])
def agregar_cliente():
    data = request.json
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute(
            'INSERT INTO clientes (nombre, cedula, celular, email, direccion) VALUES (?,?,?,?,?)',
            (data['nombre'], data['cedula'], data['celular'], data['email'], data['direccion'])
        )
        conn.commit()
        conn.close()
        return jsonify({'success': True, 'message': 'Cliente registrado correctamente'})
    except sqlite3.IntegrityError:
        return jsonify({'success': False, 'message': 'La cédula ya está registrada'}), 400
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 400

@app.route('/api/clientes/<int:id>', methods=['DELETE'])
def eliminar_cliente(id):
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute('DELETE FROM clientes WHERE id = ?', (id,))
        conn.commit()
        conn.close()
        return jsonify({'success': True, 'message': 'Cliente eliminado'})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 400

# PRODUCTOS
@app.route('/productos')
def productos():
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM productos ORDER BY nombre')
    productos = cursor.fetchall()
    conn.close()
    return render_template('productos.html', productos=productos)

@app.route('/api/productos', methods=['POST'])
def agregar_producto():
    data = request.json
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute(
            'INSERT INTO productos (nombre, descripcion, precio, stock, proveedor) VALUES (?,?,?,?,?)',
            (data['nombre'], data.get('descripcion', ''), data['precio'], data['stock'], data.get('proveedor', ''))
        )
        conn.commit()
        conn.close()
        return jsonify({'success': True, 'message': 'Producto registrado correctamente'})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 400

@app.route('/api/productos/<int:id>', methods=['DELETE'])
def eliminar_producto(id):
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute('DELETE FROM productos WHERE id = ?', (id,))
        conn.commit()
        conn.close()
        return jsonify({'success': True, 'message': 'Producto eliminado'})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 400

# VENTAS
@app.route('/ventas')
def ventas():
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM clientes ORDER BY nombre')
    clientes = cursor.fetchall()
    cursor.execute('SELECT * FROM productos ORDER BY nombre')
    productos = cursor.fetchall()
    conn.close()
    
    numero_factura = obtener_numero_factura()
    
    return render_template('ventas.html', 
                          clientes=clientes, 
                          productos=productos,
                          numero_factura=numero_factura)

@app.route('/api/ventas', methods=['POST'])
def registrar_venta():
    data = request.json
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        now = datetime.now()
        fecha = now.strftime('%Y-%m-%d')
        hora = now.strftime('%H:%M:%S')
        
        # Insertar venta
        cursor.execute(
            'INSERT INTO ventas (numero_factura, cliente_id, total, fecha, hora) VALUES (?,?,?,?,?)',
            (data['numero_factura'], data['cliente_id'], data['total'], fecha, hora)
        )
        venta_id = cursor.lastrowid
        
        # Insertar detalles de venta y actualizar stock
        for item in data['items']:
            cursor.execute(
                'INSERT INTO detalle_ventas (venta_id, producto_id, cantidad, precio, subtotal) VALUES (?,?,?,?,?)',
                (venta_id, item['producto_id'], item['cantidad'], item['precio'], item['subtotal'])
            )
            
            cursor.execute(
                'UPDATE productos SET stock = stock - ? WHERE id = ?',
                (item['cantidad'], item['producto_id'])
            )
        
        conn.commit()
        conn.close()
        return jsonify({'success': True, 'message': 'Venta registrada correctamente'})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 400

# REPORTES
@app.route('/reportes')
def reportes():
    return render_template('reportes.html')

@app.route('/api/reportes/ventas-excel')
def reporte_ventas_excel():
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        cursor.execute('''
            SELECT v.numero_factura, c.nombre, v.total, v.fecha, v.hora
            FROM ventas v
            JOIN clientes c ON v.cliente_id = c.id
            ORDER BY v.fecha DESC
        ''')
        ventas = cursor.fetchall()
        
        # Crear workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Ventas"
        
        # Encabezados
        headers = ['Factura', 'Cliente', 'Total', 'Fecha', 'Hora']
        ws.append(headers)
        
        # Formatear encabezados
        header_fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        
        # Agregar datos
        total_vendido = 0
        for venta in ventas:
            ws.append(venta)
            total_vendido += venta[2]
        
        # Ajustar columnas
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 12
        
        # Total
        ws.append([])
        ws.append(['TOTAL', '', total_vendido, '', ''])
        
        conn.close()
        
        # Guardar a bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                        as_attachment=True, download_name=f'Reporte_Ventas_{datetime.now().strftime("%Y%m%d")}.xlsx')
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 400

@app.route('/api/reportes/inventario-excel')
def reporte_inventario_excel():
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        cursor.execute('SELECT nombre, descripcion, precio, stock, proveedor FROM productos ORDER BY nombre')
        productos = cursor.fetchall()
        
        # Crear workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Inventario"
        
        # Encabezados
        headers = ['Producto', 'Descripción', 'Precio', 'Stock', 'Proveedor']
        ws.append(headers)
        
        # Formatear encabezados
        header_fill = PatternFill(start_color="2196F3", end_color="2196F3", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        
        # Agregar datos
        for producto in productos:
            ws.append(producto)
        
        # Ajustar columnas
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 20
        
        conn.close()
        
        # Guardar a bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                        as_attachment=True, download_name=f'Reporte_Inventario_{datetime.now().strftime("%Y%m%d")}.xlsx')
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 400

@app.route('/api/reportes/ventas-pdf')
def reporte_ventas_pdf():
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        cursor.execute('''
            SELECT v.numero_factura, c.nombre, v.total, v.fecha, v.hora
            FROM ventas v
            JOIN clientes c ON v.cliente_id = c.id
            ORDER BY v.fecha DESC
        ''')
        ventas = cursor.fetchall()
        conn.close()
        
        # Crear PDF
        output = io.BytesIO()
        doc = SimpleDocTemplate(output, pagesize=letter)
        elements = []
        
        # Titulo
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'], fontSize=20, textColor=colors.HexColor('#4CAF50'), spaceAfter=30)
        elements.append(Paragraph('REPORTE DE VENTAS', title_style))
        elements.append(Spacer(1, 0.2*inch))
        
        # Tabla
        data = [['Factura', 'Cliente', 'Total', 'Fecha', 'Hora']]
        total = 0
        for venta in ventas:
            data.append(list(venta))
            total += venta[2]
        
        data.append(['TOTAL', '', f'${total:,.2f}', '', ''])
        
        table = Table(data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4CAF50')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#4CAF50')),
            ('TEXTCOLOR', (0, -1), (-1, -1), colors.whitesmoke),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        elements.append(table)
        doc.build(elements)
        
        output.seek(0)
        return send_file(output, mimetype='application/pdf', as_attachment=True, download_name=f'Reporte_Ventas_{datetime.now().strftime("%Y%m%d")}.pdf')
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 400

@app.route('/api/reportes/inventario-pdf')
def reporte_inventario_pdf():
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute('SELECT nombre, descripcion, precio, stock, proveedor FROM productos ORDER BY nombre')
        productos = cursor.fetchall()
        conn.close()
        
        # Crear PDF
        output = io.BytesIO()
        doc = SimpleDocTemplate(output, pagesize=letter)
        elements = []
        
        # Titulo
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'], fontSize=20, textColor=colors.HexColor('#2196F3'), spaceAfter=30)
        elements.append(Paragraph('REPORTE DE INVENTARIO', title_style))
        elements.append(Spacer(1, 0.2*inch))
        
        # Tabla
        data = [['Producto', 'Descripción', 'Precio', 'Stock', 'Proveedor']]
        for producto in productos:
            data.append(list(producto))
        
        table = Table(data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2196F3')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        elements.append(table)
        doc.build(elements)
        
        output.seek(0)
        return send_file(output, mimetype='application/pdf', as_attachment=True, download_name=f'Reporte_Inventario_{datetime.now().strftime("%Y%m%d")}.pdf')
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 400

if __name__ == '__main__':
    init_db()
    app.run(debug=True)
